import win32com.client
import threading
from random import randint
import openpyxl




wb1 = openpyxl.load_workbook(filename='1.xlsx')
matrix1tabl = wb1.active
wb2 = openpyxl.load_workbook(filename='2.xlsx')
matrix2tabl = wb2.active
wb3 = openpyxl.load_workbook(filename='3.xlsx')
matrix3 = wb3.active

lock = threading.Lock()
def gen_new_table(mt, row, column, min, max):

    for i in range(row):
        for j in range(column):
            mt.cell(row=i+1, column=j+1).value = randint(min, max)

def gen_new_matrix(row, column, min, max):
    matrix = []
    for i in range(row):
        matrix.append([])
        for j in range(column):
            matrix[i].append(randint(min, max))
    return matrix

matrix = gen_new_matrix(100, 100, 1, 100)

gen_new_table(matrix1tabl, 100, 100, 1, 100)
wb1.save('1.xlsx')

gen_new_table(matrix2tabl, 100, 100, 1, 100)
wb2.save('2.xlsx')



def get_row(matrix, number):
    return matrix[number]

def get_column(matrix, number):
    return [i[number] for i in matrix]

matrix1 = [[item for item in row]for row in matrix1tabl.values]
lenrow1 = len(matrix1)
lencolumn1 = len(matrix1[0])

matrix2 = [[item for item in row]for row in matrix2tabl.values]
lenrow2 = len(matrix2)
lencolumn2 = len(matrix2[0])



def math(n):
    global completed

    rez = [sum(map(lambda x, y: x*y, get_row(matrix1, n), get_column(matrix2, i))) for i in range(lencolumn2)]
    for i in range(len(rez)):
        matrix3.cell(row=n+1, column=i+1).value = rez[i]
    print(rez)
    lock.acquire()
    wb3.save('3.xlsx')
    lock.release()



def main():
    for i in range(lenrow1):
        p = threading.Thread(target=math, name=f"math{i}", args=[i,])
        p.daemon = True
        p.start()


main()

inp = ''

fm = False

while inp != 'q':
    inp = input()




